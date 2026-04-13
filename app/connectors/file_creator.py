"""
Création de fichiers PDF et Excel à la demande de l'utilisateur. (TOOL-CREATE-FILES)

Fonctions exportées :
  create_pdf(title, content, username)         → dict {file_id, filename, path}
  create_excel(title, data, headers, username) → dict {file_id, filename, path}

Les fichiers sont générés dans /tmp/{uuid}.{ext} et servis via GET /download/{file_id}.
Durée de vie : les fichiers /tmp sont supprimés au redémarrage Railway (éphémère).
"""
import os
import re
import uuid
from datetime import date


def _slugify(text: str) -> str:
    """Convertit un titre en nom de fichier safe."""
    text = text.lower().strip()
    text = re.sub(r'[àáâãäå]', 'a', text)
    text = re.sub(r'[èéêë]', 'e', text)
    text = re.sub(r'[ìíîï]', 'i', text)
    text = re.sub(r'[òóôõö]', 'o', text)
    text = re.sub(r'[ùúûü]', 'u', text)
    text = re.sub(r'[ç]', 'c', text)
    text = re.sub(r'[^a-z0-9]', '_', text)
    text = re.sub(r'_+', '_', text).strip('_')
    return text[:50] or "fichier"


def create_pdf(title: str, content: str, username: str) -> dict:
    """
    Génère un PDF professionnel avec ReportLab.

    content : texte libre avec \\n pour les sauts de ligne.
               Les lignes au format "col1|col2|col3" sont rendues comme tableaux.

    Retourne {"file_id": str, "filename": str, "path": str}.
    Lève RuntimeError si reportlab n'est pas installé.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except ImportError:
        raise RuntimeError(
            "reportlab non installé. Ajoutez 'reportlab' dans requirements.txt."
        )

    file_id = str(uuid.uuid4())
    slug = _slugify(title)
    filename = f"{slug}.pdf"
    path = f"/tmp/{file_id}.pdf"

    doc = SimpleDocTemplate(
        path,
        pagesize=A4,
        rightMargin=2 * cm, leftMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()

    header_style = ParagraphStyle(
        'RayaHeader',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#6366f1'),
        spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        'RayaMeta',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=14,
    )
    body_style = ParagraphStyle(
        'RayaBody',
        parent=styles['Normal'],
        fontSize=10,
        leading=15,
        spaceAfter=6,
    )

    story = []

    # En-tête : logo texte, titre, date + auteur
    story.append(Paragraph(f"\u26a1 Raya \u2014 {title}", header_style))
    story.append(Paragraph(
        f"Généré le {date.today().strftime('%d/%m/%Y')}\u00a0\u00b7\u00a0{username}",
        meta_style,
    ))
    story.append(Spacer(1, 0.3 * cm))

    # Corps : détection automatique des blocs tableau (lignes avec |)
    lines = content.split('\n')
    i = 0
    while i < len(lines):
        line = lines[i]
        if '|' in line:
            # Collecte toutes les lignes consécutives avec | → bloc tableau
            table_lines = []
            while i < len(lines) and '|' in lines[i]:
                row = [c.strip() for c in lines[i].split('|')]
                table_lines.append(row)
                i += 1
            if table_lines:
                col_count = max(len(r) for r in table_lines)
                # Normalise la largeur de chaque ligne
                data = [r + [''] * (col_count - len(r)) for r in table_lines]
                t = Table(data, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#1e3a5f')),
                    ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
                    ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
                    ('FONTSIZE',       (0, 0), (-1, -1), 9),
                    ('GRID',           (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7fa')]),
                    ('TOPPADDING',     (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING',  (0, 0), (-1, -1), 5),
                    ('LEFTPADDING',    (0, 0), (-1, -1), 7),
                    ('RIGHTPADDING',   (0, 0), (-1, -1), 7),
                ]))
                story.append(t)
                story.append(Spacer(1, 0.4 * cm))
        elif line.strip():
            story.append(Paragraph(line.strip(), body_style))
            i += 1
        else:
            story.append(Spacer(1, 0.25 * cm))
            i += 1

    doc.build(story)

    return {
        "file_id": file_id,
        "filename": filename,
        "path": path,
    }


def create_excel(title: str, data: list, headers: list, username: str) -> dict:
    """
    Génère un fichier Excel stylisé avec openpyxl.

    headers : liste de noms de colonnes (première ligne, en-têtes).
    data    : liste de lignes, chaque ligne est une liste de valeurs.

    Retourne {"file_id": str, "filename": str, "path": str}.
    Lève RuntimeError si openpyxl n'est pas installé.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError(
            "openpyxl non installé. Ajoutez 'openpyxl' dans requirements.txt."
        )

    file_id = str(uuid.uuid4())
    slug = _slugify(title)
    filename = f"{slug}.xlsx"
    path = f"/tmp/{file_id}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel : nom de feuille limité à 31 caractères

    # ── Styles ──
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1E3A5F")  # bleu foncé Raya
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin         = Side(style='thin', color='CCCCCC')
    thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill     = PatternFill("solid", fgColor="F5F7FA")  # gris très clair

    # ── Ligne titre (row 1) ──
    title_cell = ws.cell(row=1, column=1, value=f"\u26a1 Raya \u2014 {title}")
    title_cell.font = Font(bold=True, size=13, color="6366F1")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    if headers:
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=max(len(headers), 1)
        )

    # ── En-têtes (row 2) ──
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=str(header))
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 22

    # ── Données (à partir de row 3) ──
    for row_idx, row in enumerate(data, 3):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 1:   # alternance gris sur lignes impaires (données)
                cell.fill = alt_fill

    # ── Auto-largeur des colonnes ──
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_w = len(str(header))
        for row in data:
            if col_idx - 1 < len(row):
                max_w = max(max_w, len(str(row[col_idx - 1])))
        ws.column_dimensions[col_letter].width = min(max(max_w + 4, 10), 50)

    wb.save(path)

    return {
        "file_id": file_id,
        "filename": filename,
        "path": path,
    }
